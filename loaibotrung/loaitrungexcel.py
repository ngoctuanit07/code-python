import openpyxl

def read_and_remove_duplicates(file_path):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Assume the data is in the first sheet (you can adjust this if needed)
    sheet = wb.active
    
    # Create a set to store unique rows
    unique_rows = set()
    
    # Iterate through rows and add them to the set
    for row in sheet.iter_rows():
        row_values = tuple(cell.value for cell in row)
        unique_rows.add(row_values)
    
    # Create a new workbook to store the unique rows
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    
    # Write the unique rows to the new sheet
    for row in unique_rows:
        new_sheet.append(row)
    
    # Save the new workbook
    new_wb.save("unique_data.xlsx")

# Example usage
file_path = "path/to/your/excel/file.xlsx"
read_and_remove_duplicates(file_path)
